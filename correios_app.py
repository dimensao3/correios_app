import streamlit as st
import pandas as pd
import pdfplumber
import re
import requests
import time
import io
import numpy as np
from datetime import datetime
from openpyxl.styles import PatternFill, Font

# Configuração da página
st.set_page_config(page_title="Sistema de Logística", layout="wide", page_icon="📦")

# ==========================================
# CUSTOMIZAÇÃO CSS
# ==========================================
st.markdown("""
    <style>
    .stFileUploader > div > div {
        background-color: #F4F6F9;
        border-radius: 10px;
        padding: 10px;
    }
    .stAlert {
        background-color: #E8F0FE;
        color: #001B3A;
        border-left-color: #FF7E00;
    }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# MENU LATERAL (SIDEBAR)
# ==========================================
st.sidebar.title("📦 Dimensão 3 Log")
st.sidebar.write("---")
st.sidebar.subheader("Navegação")
menu = st.sidebar.radio(
    "Escolha a Ferramenta:", 
    ["📄 Extrator de Rastreios", "📍 Processador de Planilhas"]
)

st.sidebar.write("---")
st.sidebar.caption("Desenvolvido para automatização de envios e geração de layouts para os Correios.")

# ==========================================
# FERRAMENTA 1: EXTRATOR DE PDFs
# ==========================================
if menu == "📄 Extrator de Rastreios":
    st.title("📄 Extrair Códigos de Rastreio (PDF)")
    st.write("Faça o upload dos PDFs gerados pelos Correios para extrair os nomes e rastreamentos.")
    
    st.markdown("<br>", unsafe_allow_html=True) 
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        arquivos_pdf = st.file_uploader("Selecione os arquivos PDF", type=["pdf"], accept_multiple_files=True)
        
        if arquivos_pdf and st.button("🚀 Processar PDFs agora"):
            dados_extraidos = []
            padrao_rastreio = re.compile(r'(\S{2}\d{9}BR)', re.IGNORECASE)
            
            with st.spinner("Lendo PDFs... Isso pode levar alguns segundos."):
                for arquivo in arquivos_pdf:
                    try:
                        with pdfplumber.open(arquivo) as pdf:
                            for pagina in pdf.pages:
                                texto = pagina.extract_text()
                                if not texto: continue
                                linhas = texto.split('\n')
                                for i, linha in enumerate(linhas):
                                    match = padrao_rastreio.search(linha)
                                    if match:
                                        codigo_rastreio = match.group(1).upper()
                                        resto_linha = linha.replace(match.group(0), '').strip()
                                        nome_pintor = re.split(r'\s\d{4,}', resto_linha)[0].strip()
                                        if not re.search(r'[A-Za-z]', nome_pintor):
                                            for offset in [1, 2, 3]:
                                                if (i + offset) < len(linhas):
                                                    linha_abaixo = linhas[i + offset].strip()
                                                    if re.search(r'[A-Za-z]', linha_abaixo) and not padrao_rastreio.search(linha_abaixo):
                                                        nome_pintor = re.split(r'\s\d{4,}', linha_abaixo)[0].strip()
                                                        break
                                        nome_pintor = re.sub(r'\b(PAC|SEDEX).*', '', nome_pintor, flags=re.IGNORECASE).strip()
                                        if nome_pintor and "Código do objeto" not in nome_pintor:
                                            dados_extraidos.append({'Nome do Pintor': nome_pintor, 'Código de Rastreio': codigo_rastreio})
                    except Exception as e:
                        st.error(f"Erro ao processar {arquivo.name}: {e}")

            if dados_extraidos:
                df_rastreios = pd.DataFrame(dados_extraidos).astype(str)
                
                st.success("Extração concluída com sucesso!")
                st.metric(label="Total de Registros Extraídos", value=f"{len(df_rastreios)} envios")
                
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_rastreios.to_excel(writer, index=False, sheet_name='Rastreios')
                    ws = writer.sheets['Rastreios']
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.number_format = '@'
                            
                st.download_button(
                    label="📥 Baixar Planilha de Rastreios (.xlsx)", 
                    data=buffer.getvalue(), 
                    file_name=f"Rastreios_{datetime.now().strftime('%d-%m_%H-%M')}.xlsx",
                    use_container_width=True
                )

# ==========================================
# FERRAMENTA 2: PROCESSADOR PARA IMPORTAÇÃO
# ==========================================
elif menu == "📍 Processador de Planilhas":
    st.title("📍 Processador para Importação dos Correios")
    st.write("Gere a planilha de postagem final formatada a partir das suas bases de envio.")
    
    st.info("""
    **📌 Tabela Auxiliar de Serviços e Formatos (Para preenchimento no Excel)**
    * **PAC:** 03298 | **SEDEX:** 03220 | **ENVELOPE:** 1 | **CAIXA:** 2 | **CILINDRO:** 3
    """)

    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.subheader("1. Tipo de Envio")
        modo_envio = st.radio("Selecione a lógica de agrupamento:", [
            "Envios de Resgates (Agrupar ID)", 
            "Envios Padrão / Avulsas (Múltiplas colunas)"
        ])
        
    with col2:
        st.subheader("2. Arquivo Base")
        arquivo_base = st.file_uploader("Selecione a Planilha (.csv ou .xlsx)", type=["csv", "xlsx"])
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    if arquivo_base and st.button("🚀 Processar Planilha de Importação", use_container_width=True):
        
        medidas_caixas = {1:(44,32,27), 2:(50,34,32), 3:(34,34,16), 4:(32,22,28), 5:(24,24,14), 6:(44,33,15), 7:(55,36,45), 8:(24,13,16)}

        def identificar_caixa(valor):
            val_str = str(valor).strip().lower()
            match = re.search(r'(?:cx|caixa)\s*0*(\d+)', val_str)
            return int(match.group(1)) if match else None

        def buscar_cep(cep):
            if not cep or len(str(cep)) != 8: return None
            try:
                r = requests.get(f"https://viacep.com.br/ws/{cep}/json/", timeout=5)
                if r.status_code == 200 and "erro" not in r.json():
                    return {'logradouro': r.json().get('logradouro', ''), 'bairro': r.json().get('bairro', ''), 'localidade': r.json().get('localidade', ''), 'uf': r.json().get('uf', '')}
            except: pass
            return None

        # FUNÇÃO CORRIGIDA: Agora remove o ".0" antes de extrair os números
        def limpa_numero(valor, tamanho):
            txt = str(valor).strip()
            txt = re.sub(r'\.0$', '', txt)  # Corta o ".0" fantasma do Excel
            txt = re.sub(r'\D', '', txt)    # Remove os não-números (ex: traços)
            return txt.zfill(tamanho) if txt and txt != '0' else ""

        with st.spinner("Processando endereços e organizando caixas..."):
            df = pd.read_csv(arquivo_base) if arquivo_base.name.endswith('.csv') else pd.read_excel(arquivo_base)
            df.columns = df.columns.str.strip().str.upper()

            for idx, row in df.iterrows():
                caixa_id = None
                for col in ['ALTURA', 'LARGURA', 'COMPRIMENTO']:
                    if col in df.columns:
                        res = identificar_caixa(row[col])
                        if res in medidas_caixas: caixa_id = res; break
                if caixa_id:
                    df.at[idx, 'ALTURA'], df.at[idx, 'LARGURA'], df.at[idx, 'COMPRIMENTO'] = medidas_caixas[caixa_id]

            if 'CPF' in df.columns: df['CPF'] = df['CPF'].apply(lambda x: limpa_numero(x, 11))
            if 'CEP' in df.columns: df['CEP'] = df['CEP'].apply(lambda x: limpa_numero(x, 8))
            df = df[df['CEP'] != ""]

            for c in ['PESO', 'ALTURA', 'LARGURA', 'COMPRIMENTO']:
                if c in df.columns:
                    df[c] = df[c].fillna("").astype(str).replace(r'\.0$', '', regex=True).replace('nan', '')

            max_itens = 0
            if "Resgates" in modo_envio:
                df = df.dropna(subset=['ID PINTOR'])
                agg_funcs = {'CPF':'first','PINTOR':'first','CEP':'first','PESO':'max','ALTURA':'max','LARGURA':'max','COMPRIMENTO':'max'}
                for c in ['NUMERO', 'COMPLEMENTO']: 
                    if c in df.columns: agg_funcs[c] = 'first'
                df_fixo = df.groupby('ID PINTOR').agg(agg_funcs).reset_index()
                df_fixo['OBSERVACAO'] = df_fixo['ID PINTOR']
                df['Material_Idx'] = df.groupby('ID PINTOR').cumcount() + 1
                max_itens = df['Material_Idx'].max()
                
                df_pivoted = df.pivot(index='ID PINTOR', columns='Material_Idx', values=['MATERIAL', 'QUANTIDADE', 'VALOR' if 'VALOR' in df.columns else 'MATERIAL'])
                df_materiais = pd.DataFrame(index=df_pivoted.index)
                for i in range(1, max_itens + 1):
                    df_materiais[f'DeclaracaoConteudoConteudo{i}'] = df_pivoted[('MATERIAL', i)]
                    df_materiais[f'DeclaracaoConteudoQuantidade{i}'] = df_pivoted[('QUANTIDADE', i)]
                    df_materiais[f'DeclaracaoConteudoValor{i}'] = df_pivoted[('VALOR', i)] if 'VALOR' in df.columns else ""
                df_base = pd.merge(df_fixo, df_materiais, on='ID PINTOR', how='left')
            else:
                df['ID_UNICO'] = range(len(df))
                df_fixo = df.set_index('ID_UNICO')
                mat_cols = [c for c in df.columns if re.match(r'MATERIAL\d*', c)]
                df_materiais = pd.DataFrame(index=df_fixo.index)
                for col in mat_cols:
                    n = re.search(r'\d+', col).group() if re.search(r'\d+', col) else "1"
                    idx = int(n)
                    max_itens = max(max_itens, idx)
                    df_materiais[f'DeclaracaoConteudoConteudo{idx}'] = df_fixo[col]
                    df_materiais[f'DeclaracaoConteudoQuantidade{idx}'] = df_fixo.get(f'QUANTIDADE{n}', "")
                    df_materiais[f'DeclaracaoConteudoValor{idx}'] = df_fixo.get(f'VALOR{n}', "")
                df_base = pd.merge(df_fixo, df_materiais, left_index=True, right_index=True)

            ceps_unicos = df_base['CEP'].unique()
            dic_ceps = {cep: buscar_cep(cep) for cep in ceps_unicos}

            cols_correios = ["sequencial", "cpfCnpjRemetente", "documentoEstrangeiroRemetente", "nomeRemetente", "dddTelefoneRemetente", "telefoneRemetente", "dddCelularRemetente", "celularRemetente", "emailRemetente", "observacaoRemetente", "cepRemetente", "logradouroRemetente", "numeroRemetente", "complementoRemetente", "bairroRemetente", "cidadeRemetente", "ufRemetente", "cpfCnpjDestinatario", "documentoEstrangeiroDestinatario", "nomeDestinatario", "dddTelefoneDestinatario", "telefoneDestinatario", "dddCelularDestinatario", "celularDestinatario", "emailDestinatario", "observacaoDestinatario", "cepDestinatario", "logradouroDestinatario", "numeroDestinatario", "complementoDestinatario", "bairroDestinatario", "cidadeDestinatario", "ufDestinatario", "codigoServico", "dataPrevistaPostagem", "prazoPostagem", "logisticaReversa", "dataValidadeLogReversa", "codigoServicoAdicionalValorDeclarado", "valorDeclarado", "codigoServicoAdicionalEntregaVizinho", "orientacaoEntregaVizinho", "codigoServicoAdicional1", "codigoServicoAdicional2", "codigoServicoAdicional3", "pesoInformado", "codigoFormatoObjetoInformado", "alturaInformada", "larguraInformada", "comprimentoInformado", "diametroInformado", "cienteObjetoNaoProibido", "observacao", "numeroNotaFiscal", "chaveNFe", "rfidObjeto"]
            for i in range(1, max(10, max_itens) + 1): cols_correios += [f'DeclaracaoConteudoConteudo{i}', f'DeclaracaoConteudoQuantidade{i}', f'DeclaracaoConteudoValor{i}']
            cols_correios.append("codigoObjetoIda")

            df_final = pd.DataFrame(columns=cols_correios)
            df_final['sequencial'] = range(1, len(df_base) + 1)
            df_final[['cpfCnpjRemetente','nomeRemetente','cepRemetente','logradouroRemetente','numeroRemetente','bairroRemetente','cidadeRemetente','ufRemetente','cienteObjetoNaoProibido']] = ['03469994000188','Dimensao 3 Log','09930450','Avenida paranapanema','614','Taboão','São Paulo','SP','1']
            
            df_final['logisticaReversa'] = 'N'
            df_final['cpfCnpjDestinatario'] = df_base['CPF'].values
            df_final['nomeDestinatario'] = df_base['PINTOR'].values
            df_final['cepDestinatario'] = df_base['CEP'].values
            df_final['logradouroDestinatario'] = df_base['CEP'].map(lambda x: dic_ceps.get(x,{}).get('logradouro') if dic_ceps.get(x) else "NÃO ENCONTRADO")
            df_final['bairroDestinatario'] = df_base['CEP'].map(lambda x: dic_ceps.get(x,{}).get('bairro','') if dic_ceps.get(x) else "")
            df_final['cidadeDestinatario'] = df_base['CEP'].map(lambda x: dic_ceps.get(x,{}).get('localidade','') if dic_ceps.get(x) else "")
            df_final['ufDestinatario'] = df_base['CEP'].map(lambda x: dic_ceps.get(x,{}).get('uf','') if dic_ceps.get(x) else "")
            df_final['numeroDestinatario'] = df_base.get('NUMERO', "").values
            df_final['complementoDestinatario'] = df_base.get('COMPLEMENTO', "").values
            df_final['pesoInformado'] = df_base['PESO'].values
            df_final['alturaInformada'] = df_base['ALTURA'].values
            df_final['larguraInformada'] = df_base['LARGURA'].values
            df_final['comprimentoInformado'] = df_base['COMPRIMENTO'].values
            df_final['observacao'] = df_base.get('OBSERVACAO', "").values

            for i in range(1, max_itens + 1):
                df_final[f'DeclaracaoConteudoConteudo{i}'] = df_base[f'DeclaracaoConteudoConteudo{i}'].values
                df_final[f'DeclaracaoConteudoQuantidade{i}'] = df_base[f'DeclaracaoConteudoQuantidade{i}'].values
                df_final[f'DeclaracaoConteudoValor{i}'] = df_base[f'DeclaracaoConteudoValor{i}'].values

            df_final = df_final.fillna("").astype(str).replace(r'\.0$', '', regex=True)

            dados_erros = []
            
            for i, row in df_final.iterrows():
                faltando_endereco = []
                if not row['numeroDestinatario'] or row['numeroDestinatario'].strip() == "": faltando_endereco.append("Número")
                if not row['logradouroDestinatario'] or row['logradouroDestinatario'] == "NÃO ENCONTRADO": faltando_endereco.append("Rua")
                if not row['bairroDestinatario'] or row['bairroDestinatario'].strip() == "": faltando_endereco.append("Bairro")
                if not row['cidadeDestinatario'] or row['cidadeDestinatario'].strip() == "": faltando_endereco.append("Cidade")
                
                if faltando_endereco:
                    dados_erros.append({"Linha (Sequencial)": i + 1, "ID Pintor": row['observacao'] if "Resgates" in modo_envio else "", "Nome": row['nomeDestinatario'], "Detalhe do Erro": f"Falta preencher: {', '.join(faltando_endereco)}."})
                
                try:
                    if float(row['alturaInformada'] or 0) > 100 or float(row['larguraInformada'] or 0) > 100 or float(row['comprimentoInformado'] or 0) > 100:
                        dados_erros.append({"Linha (Sequencial)": i + 1, "ID Pintor": row['observacao'] if "Resgates" in modo_envio else "", "Nome": row['nomeDestinatario'], "Detalhe do Erro": "Medida acima de 100cm detectada."})
                except: pass
            
            st.write("---")
            st.subheader("📋 Relatório de Auditoria")
            
            if dados_erros:
                df_erros = pd.DataFrame(dados_erros)
                if "Resgates" not in modo_envio: df_erros = df_erros.drop(columns=["ID Pintor"])
                
                for idx, erro in df_erros.head(10).iterrows():
                    detalhe_linha = f"**Linha {erro['Linha (Sequencial)']} ({erro['Nome']}):** {erro['Detalhe do Erro']}"
                    if "Resgates" in modo_envio and erro['ID Pintor']:
                        detalhe_linha = f"**Linha {erro['Linha (Sequencial)']} - ID {erro['ID Pintor']} ({erro['Nome']}):** {erro['Detalhe do Erro']}"
                    st.error("❌ " + detalhe_linha)
                
                if len(df_erros) > 10:
                    st.warning(f"⚠️ Existem mais {len(df_erros) - 10} erros não listados aqui. Baixe o relatório completo.")
                    buffer_erros = io.BytesIO()
                    with pd.ExcelWriter(buffer_erros, engine='openpyxl') as writer:
                        df_erros.to_excel(writer, index=False, sheet_name='Erros')
                    st.download_button(label="⚠️ Baixar Relatório de Erros", data=buffer_erros.getvalue(), file_name=f"Relatorio_Erros_{datetime.now().strftime('%d-%m_%H-%M')}.xlsx")
            else:
                st.success("✅ Nenhum erro crítico detectado nos dados!")

            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name='Importacao')
                ws = writer.sheets['Importacao']
                
                fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                fill_laranja = PatternFill(start_color="FF7E00", end_color="FF7E00", fill_type="solid")
                fill_verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                fill_azul = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
                fill_black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                font_white = Font(color="FFFFFF")
                
                colunas_cabecalho_pretas = [
                    "cpfCnpjDestinatario", "nomeDestinatario", "cepDestinatario", 
                    "logradouroDestinatario", "numeroDestinatario", "bairroDestinatario", 
                    "cidadeDestinatario", "ufDestinatario", "codigoServico", 
                    "logisticaReversa", "pesoInformado", "codigoFormatoObjetoInformado", 
                    "alturaInformada", "larguraInformada", "comprimentoInformado"
                ]
                
                colunas_laranja = [
                    "cpfCnpjRemetente", "nomeRemetente", "cepRemetente", "logradouroRemetente", 
                    "numeroRemetente", "bairroRemetente", "cidadeRemetente", "ufRemetente",
                    "cienteObjetoNaoProibido"
                ]

                for row_idx, row in enumerate(ws.iter_rows(), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        cell.number_format = '@'
                        if row_idx == 1:
                            col_name = cell.value
                            if col_name == "sequencial": cell.fill = fill_verde
                            elif col_name in colunas_cabecalho_pretas: 
                                cell.fill = fill_black
                                cell.font = font_white
                            elif col_name in colunas_laranja: cell.fill = fill_laranja
                            elif col_name and str(col_name).startswith("DeclaracaoConteudo"): cell.fill = fill_azul
                            elif col_name == "observacao":
                                if "Resgates" in modo_envio:
                                    cell.fill = fill_black
                                    cell.font = font_white
                                else:
                                    cell.fill = fill_azul

                for col_name in ["codigoServico", "codigoFormatoObjetoInformado"]:
                    col_idx = df_final.columns.get_loc(col_name) + 1
                    for r in range(2, len(df_final) + 2): ws.cell(r, col_idx).fill = fill_yellow
                
                colunas_auditoria = ["numeroDestinatario", "logradouroDestinatario", "bairroDestinatario", "cidadeDestinatario"]
                for col_name in colunas_auditoria:
                    col_idx = df_final.columns.get_loc(col_name) + 1
                    for r in range(2, len(df_final) + 2):
                        valor_cel = str(ws.cell(r, col_idx).value).strip().upper()
                        if not valor_cel or valor_cel == "NONE" or valor_cel == "NÃO ENCONTRADO":
                            ws.cell(r, col_idx).fill = fill_red

            st.write("---")
            st.download_button(
                label="📥 Baixar Planilha para Importação Pronta", 
                data=buffer.getvalue(), 
                file_name=f"Importacao_Correios_{datetime.now().strftime('%d-%m_%H-%M')}.xlsx",
                use_container_width=True
            )
